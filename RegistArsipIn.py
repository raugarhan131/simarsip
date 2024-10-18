import fitz  # PyMuPDF
import re
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import os
import tkinter as tk
#from tkinter import filedialog, ttk, messagebox
import shutil
import PySimpleGUI as sg
import pytesseract
from pdf2image import convert_from_path
from PIL import Image
import streamlit as st
import io
import time
import datetime
import pandas as pd
import streamlit as st
from st_aggrid import AgGrid, GridOptionsBuilder
import sqlite3
import explor

# Fungsi untuk mengonversi PDF ke gambar dan menjalankan OCR
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Fungsi untuk mengambil data Klasifikasi Primer
def get_Primer():
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  cursor.execute("SELECT id, Kode, Nama, ParentId, ketSKKA, StatusRetensi, Deskripsi  FROM KA WHERE ParentId=0")  # Primer memiliki ParentId=0
  primer = cursor.fetchall()
  conn.close()
  return primer

# Fungsi untuk mengambil data Klasifikasi Sekunder berdasarkan primer_id
def get_Sekunder(primer_id):
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  cursor.execute("SELECT id, Kode, Nama, ParentId, ketSKKA, StatusRetensi, Deskripsi  FROM KA WHERE ParentId=1 AND Kode like ?", (f"{primer_id}%",))
  sekunder = cursor.fetchall()
  conn.close()
  return sekunder

# Fungsi untuk mengambil data Klasifikasi Sekunder berdasarkan primer_id
def get_Tersier(sekunder_id):
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  cursor.execute("SELECT id, Kode, Nama, ParentId, ketSKKA, StatusRetensi, Deskripsi  FROM KA WHERE ParentId=2 AND Kode like ?", (f"{sekunder_id}%",))
  sekunder = cursor.fetchall()
  conn.close()
  return sekunder

def get_aktif():
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  query="""
  SELECT  Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInaktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi FROM ArsipIn WHERE Status='Aktif' 
  order by Tahun, NomorSurat, TglSurat, Kode
  """
  cursor.execute(query)  
  aktif = cursor.fetchall()
  conn.close()
  return aktif

def edit_retensi():
    conn = sqlite3.connect('simarsip.db')
    cursor = conn.cursor()

    query = """
    UPDATE ArsipOut
    SET Retensi = (SELECT KA.StatusRetensi FROM KA WHERE KA.Kode = ArsipOut.Kode),
        KetSKKA = (SELECT KA.KetSKKA FROM KA WHERE KA.Kode = ArsipOut.Kode),
        RetAktif = (SELECT KA.RetAktif FROM KA WHERE KA.Kode = ArsipOut.Kode),
        Retinaktif = (SELECT KA.RetInaktif FROM KA WHERE KA.Kode = ArsipOut.Kode)
    WHERE EXISTS (SELECT 1 FROM KA WHERE KA.Kode = ArsipOut.Kode);
    """
    cursor.execute(query)
    conn.commit()

    query = """
    UPDATE ArsipIn
    SET Retensi = (SELECT KA.StatusRetensi FROM KA WHERE KA.Kode = ArsipIn.Kode),
        KetSKKA = (SELECT KA.KetSKKA FROM KA WHERE KA.Kode = ArsipIn.Kode),
        RetAktif = (SELECT KA.RetAktif FROM KA WHERE KA.Kode = ArsipIn.Kode),
        Retinaktif = (SELECT KA.RetInaktif FROM KA WHERE KA.Kode = ArsipIn.Kode)
    WHERE EXISTS (SELECT 1 FROM KA WHERE KA.Kode = ArsipIn.Kode);
    """
    cursor.execute(query)
    conn.commit()

    conn.close()

    return 

def convert_df_to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    processed_data = output.getvalue()
    return processed_data

# Fungsi untuk memasukkan data arsip ke tabel 
def insert_arsip(Pengguna_id, UnitKerja_id, KA_id, Berkas_id, NomorSurat, TglSurat, Kode, Jenis, Hal, IsiSurat, Lampiran, Ttd, Direktori, NamaFile, Status, KetSKKA, Retensi, TglEntri, Lokasi, Pengguna, tahun):
  try:
    conn = sqlite3.connect('simarsip.db')
    c = conn.cursor()
    c.execute('''
      INSERT INTO ArsipIn (Pengguna_id, UnitKerja_id, KA_id, Berkas_id, NomorSurat, TglSurat, Kode, Jenis, Hal, IsiSurat, Lampiran, Ttd, Direktori, NamaFile, Status, KetSKKA, Retensi, TglEntri, Lokasi, Pengguna, Tahun) 
      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
      ''', (Pengguna_id, UnitKerja_id, KA_id, Berkas_id, NomorSurat, TglSurat, Kode, Jenis, Hal, IsiSurat, Lampiran, Ttd, Direktori, NamaFile, Status, KetSKKA, Retensi, TglEntri, Lokasi, Pengguna, tahun))
    conn.commit()
    st.success("File berhasil diunggah dan disimpan")
  except sqlite3.Error as e:
    st.error(f"Error saat menyimpan data: {e}")
  finally:
    conn.close()

def convert_pdfNonscan_text(pdf_document, namafile): #1. Pdf Non Scan
  file_name = pdf_document #os.path.basename(pdf_document)
  jum_hal=0
  jenis_pdf =""

  doc = fitz.open(stream=pdf_document.read(), filetype="pdf")
  page = doc.load_page(0)  # Load halaman pertama
  jum_hal = doc.page_count
  text = page.get_text("text")

  doc.close()
  attributes = extract_attributes(text)
  jenis_pdf = "Non Scan"

  tahun = attributes['tahun']
  tanggal_surat = attributes['tanggal_surat']
  nomor_surat = attributes['nomor_surat']
  hal_surat = attributes['hal_surat']
  kode_surat = attributes['kode_surat']
  retensi = attributes['retensi']
  skka = attributes['skka']

  # Periksa apakah variabel berisi None atau kode_surat berisi "xxxxxxxxxxxxxxxxx"
  if None in [hal_surat, nomor_surat, tanggal_surat] or kode_surat == "xxxxxxxxxxxxxxxxx":
    st.error("Data surat tidak terbaca, silakan input manual.")

  else:       
    # Simpan file yang diunggah ke sistem lokal
    if pdf_document is not None:
      # Menyimpan file di direktori lokal
      upload_folder = "uploaded_files/"
      if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)
      save_path = os.path.join(upload_folder, pdf_document.name)
          
      # Menyimpan file di server lokal
      with open(save_path, "wb") as f:
        f.write(pdf_document.getbuffer())

    berkas(kode_surat, save_path, namafile, tahun, nomor_surat, tanggal_surat,  hal_surat,  retensi, skka, jum_hal)

  return

def convert_pdfscan_text1(pdf_document, namafile): #2. Pdf Scan - 1
  # Convert PDF to images
  pages = convert_from_path(namafile, 300)
  jum_hal = 0
  # Iterate through all the pages
  for page_number, page_data in enumerate(pages):
    # Save the page as an image file
    image_path = f'page_{page_number + 1}.jpg'
    page_data.save(image_path, 'JPEG')
    jum_hal+=1
    # Use Tesseract to extract text from the image
    text = pytesseract.image_to_string(Image.open(image_path))

  attributes = extract_attributes(text)
  jenis_pdf = "Non Scan"

  tahun = attributes['tahun']
  tanggal_surat = attributes['tanggal_surat']
  nomor_surat = attributes['nomor_surat']
  hal_surat = attributes['hal_surat']
  kode_surat = attributes['kode_surat']
  retensi = attributes['retensi']
  skka = attributes['skka']

  # Periksa apakah variabel berisi None atau kode_surat berisi "xxxxxxxxxxxxxxxxx"
  if None in [hal_surat, nomor_surat, tanggal_surat] or kode_surat == "xxxxxxxxxxxxxxxxx":
    st.warning("Beberapa data surat tidak terbaca, silakan input manual.")
    
    col1 = st.columns(1)
    entri_button = col1.button("Entri Manual")
    if entri_button:
    # Tampilkan form input untuk meminta data manual dari user
      with st.form("input_data_form"):
        #KA_id = st.number_input("KA ID", min_value=1, step=1, format="%d")
        #Berkas_id = st.number_input("Unit Kerja ID", min_value=1, step=1, format="%d")
        no = st.text_input("Nomor Surat", min_value=1, step=1)
        kodeunit = st.text_input("Kode Unit Kerja Pengirim")
        primer = st.text_input("Kode Klasifikasi Primer")
        sekunder = st.text_input("Kode Klasifikasi Sekunder")
        tersier = st.text_input("Kode Klasifikasi tersier")
        strtanggal_surat = st.date_input("Tanggal Surat", value=datetime.date.today())
        retensi = st.text_input("Jenis Retensi (M/P/DK)")
        skka = st.text_input("Kode SKKA (B/T)")
        tahun = strtanggal_surat.year
        hal_surat = st.text_input("Perihal Surat")
        #IsiSurat = st.text_input("Isi/Rangkuman Surat")
        #Ttd = st.text_input("Pengirim")
        #Lampiran = st.text_input("Jumlah Lampiran", min_value=0, step=1)
        kode_surat = f"{primer}.{sekunder}.{tersier}"
        nomor_surat = f"{no}/{kodeunit}/{primer}.{sekunder}.{tersier}/{retensi}/{skka}/{tahun}"
        tanggal_surat = strtanggal_surat.strftime("%d-%b-%Y")
        submitadd = st.form_submit_button("Simpan Data")        

        if submitadd:
          st.write("data sudah diisi")

  # Simpan file yang diunggah ke sistem lokal
  if pdf_document is not None:
    # Menyimpan file di direktori lokal
    upload_folder = "uploaded_files/"
    if not os.path.exists(upload_folder):
      os.makedirs(upload_folder)
    save_path = os.path.join(upload_folder, pdf_document.name)
        
    # Menyimpan file di server lokal
    with open(save_path, "wb") as f:
      f.write(pdf_document.getbuffer())

  berkas(kode_surat, save_path, namafile, tahun, nomor_surat, tanggal_surat,  hal_surat,  retensi, skka, jum_hal)

  return 

def convert_pdfscan_text2(pdf_document, namafile): #3. Pdf Scan -2
  poppler_path = r'C:\\poppler\\Library\\bin' 
  text=""
  
  if pdf_document is not None:
    # Simpan file PDF sementara
    pdf_document.seek(0)
    doc = fitz.open(stream=pdf_document.read(), filetype="pdf")
    ocr_text = ""
    jum_hal = 0
    # Loop setiap halaman PDF
    for page_num in range(len(doc)):
      #st.write(f"Memproses halaman {page_num + 1}...")
      jum_hal+=1
      # Konversi halaman PDF ke gambar
      page = doc.load_page(page_num)  # Memuat halaman
      pix = page.get_pixmap()  # Konversi ke Pixmap (gambar)
        
      # Simpan gambar sebagai byte array
      img_bytes = pix.tobytes("png")
        
      # Konversi byte array menjadi image PIL
      image = Image.open(io.BytesIO(img_bytes))
        
      # Lakukan OCR pada gambar halaman menggunakan Pytesseract
      page_text = pytesseract.image_to_string(image)
        
      # Tambahkan teks dari halaman ke variabel ocr_text
      ocr_text += f"\n\nHalaman {page_num + 1}:\n"
      ocr_text += merge_lines(page_text)
    
    # Tampilkan hasil OCR
    #st.text_area("Teks Hasil OCR", ocr_text, height=300)
    
    # Opsi untuk mengunduh hasil OCR sebagai file teks
    if ocr_text:
        ocr_file = io.StringIO(ocr_text)
        #st.download_button(label="Unduh Hasil OCR", data=ocr_file.getvalue(), file_name="ocr_output.txt", mime="text/plain")

    attributes = extract_attributes(merge_lines(ocr_text))
    jenis_pdf = "Non Scan"

    tahun = attributes['tahun']
    tanggal_surat = attributes['tanggal_surat']
    nomor_surat = attributes['nomor_surat']
    hal_surat = attributes['hal_surat']
    kode_surat = attributes['kode_surat']
    retensi = attributes['retensi']
    skka = attributes['skka']

  # Periksa apakah variabel berisi None atau kode_surat berisi "xxxxxxxxxxxxxxxxx"
  if None in [hal_surat, nomor_surat, tanggal_surat] or kode_surat == "xxxxxxxxxxxxxxxxx":
    st.warning("Beberapa data surat tidak terbaca, silakan input manual.")
    
    col1 = st.columns(1)
    entri_button = col1.button("Entri Manual")
    if entri_button:
    # Tampilkan form input untuk meminta data manual dari user
      with st.form("input_data_form"):
        #Pengguna_id = st.number_input("Pengguna ID", min_value=1, step=1, format="%d")
        #UnitKerja_id = st.number_input("Unit Kerja ID", min_value=1, step=1, format="%d")
        #KA_id = st.number_input("KA ID", min_value=1, step=1, format="%d")
        #Berkas_id = st.number_input("Unit Kerja ID", min_value=1, step=1, format="%d")
        no = st.text_input("Nomor Surat", min_value=1, step=1)
        kodeunit = st.text_input("Kode Unit Kerja Pengirim")
        primer = st.text_input("Kode Klasifikasi Primer")
        sekunder = st.text_input("Kode Klasifikasi Sekunder")
        tersier = st.text_input("Kode Klasifikasi tersier")
        strtanggal_surat = st.date_input("Tanggal Surat", value=datetime.date.today())
        retensi = st.text_input("Jenis Retensi (M/P/DK)")
        skka = st.text_input("Kode SKKA (B/T)")
        tahun = strtanggal_surat.year
        hal_surat = st.text_input("Perihal Surat")
        #IsiSurat = st.text_input("Isi/Rangkuman Surat")
        #Ttd = st.text_input("Pengirim")
        #Lampiran = st.text_input("Jumlah Lampiran", min_value=0, step=1)
        kode_surat = f"{primer}.{sekunder}.{tersier}"
        nomor_surat = f"{no}/{kodeunit}/{primer}.{sekunder}.{tersier}/{retensi}/{skka}/{tahun}"
        tanggal_surat = strtanggal_surat.strftime("%d-%b-%Y")
        submitadd = st.form_submit_button("Simpan Data")        

        if submitadd:
          st.write("data sudah diisi")

    # Simpan file yang diunggah ke sistem lokal
    if pdf_document is not None:
      upload_folder = "uploaded_files/"
      if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)
      save_path = os.path.join(upload_folder, pdf_document.name)
          
      with open(save_path, "wb") as f:
        f.write(pdf_document.getbuffer())

    berkas(kode_surat, save_path, namafile, tahun, nomor_surat, tanggal_surat,  hal_surat,  retensi, skka, jum_hal)    
  return 

def berkas(kode_surat, save_path, namafile, tahun, nomor_surat, tanggal_surat,  hal_surat,  retensi, skka, jum_hal):
  jum_Kode = len(kode_surat)
  root_path="Berkas/"
  if not os.path.exists(root_path):
    os.makedirs(root_path, exist_ok=True)
  if jum_Kode==2 :
    primer_folder = root_path + kode_surat[0:2]
    if not os.path.exists(primer_folder):
      os.makedirs(primer_folder, exist_ok=True)
    shutil.copy(save_path, primer_folder) #+ "/" + file_name)
  elif jum_Kode==5:
    primer_folder = root_path + kode_surat[0:2] 
    if not os.path.exists(primer_folder):
      os.makedirs(primer_folder, exist_ok=True)

    sekunder_folder = primer_folder + "/" + kode_surat[0:5]
    if not os.path.exists(sekunder_folder):
      os.makedirs(sekunder_folder, exist_ok=True)
    shutil.copy(save_path, sekunder_folder) #+ "/") # + file_name)
  elif jum_Kode==8:
    primer_folder = root_path + kode_surat[0:2] 
    if not os.path.exists(primer_folder):
      os.makedirs(primer_folder, exist_ok=True)

    sekunder_folder = primer_folder + "/" + kode_surat[0:5]
    if not os.path.exists(sekunder_folder):
      os.makedirs(sekunder_folder, exist_ok=True)                

    tersier_folder = sekunder_folder + "/" + kode_surat[0:8] 
    if not os.path.exists(tersier_folder):
      os.makedirs(tersier_folder, exist_ok=True)
    shutil.copy(save_path, tersier_folder) 

  os.remove(save_path)
  file_link=" "
  if jum_Kode==2:
    file_link = primer_folder +  "/" + namafile
  elif jum_Kode==5:
    file_link = sekunder_folder +  "/" + namafile
  elif jum_Kode==8:
    file_link = tersier_folder + "/" + namafile

  # simpan data ke excel
  file_path = 'DaftarArsipMasuk.xlsx'
  if os.path.exists(file_path): #jika file sudah ada
    if check_if_file_is_open(file_path) == True: #jika file tdk sdg terbuka
      workbook = load_workbook(file_path)
      worksheet = workbook.active
        
      # Menambahkan baris baru
      data = [tahun, nomor_surat, kode_surat, tanggal_surat,  hal_surat,  retensi, skka, file_link, jum_hal ]
      worksheet.append(data)
      last_row = worksheet.max_row

      worksheet.cell(row=last_row, column=8).hyperlink = file_link
      worksheet.cell(row=last_row, column=8).value = file_link  # Text yang akan tampil
      worksheet.cell(row=last_row, column=8).style = "Hyperlink"    # Gaya hyperlink

      workbook.save(file_path)
    else:
      st.warning("File DaftarArsipMasuk.xlsx sedang terbuka, silakan tutup terlebih dahulu.")
      st.stop()

  else: #jika file blm ada
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    # Header
    header = ['Tahun', 'No Surat', 'Kode', 'Tgl Surat', 'Hal Surat', 'Retensi', 'SKKA', 'Lokasi Arsip', 'Lampiran']
    worksheet.write_row(0, 0, header)
    data = [tahun, nomor_surat, kode_surat, tanggal_surat, hal_surat, retensi, skka, file_link, jum_hal]
    worksheet.write_row(1, 0, data)
    worksheet.write_url(1, 7, file_link, string=file_link)  # Hyperlink di kolom 8
      
  workbook.close()
  Pengguna_id = st.session_state['id']
  UnitKerja_id = st.session_state['Unit_id']
  Pengguna = st.session_state['namauser']
  KA_id=1
  Berkas_id=1
  NomorSurat=nomor_surat
  TglSurat=tanggal_surat
  Kode=kode_surat
  Jenis="Aktif"
  Hal=hal_surat
  IsiSurat=hal_surat
  Lampiran=jum_hal
  Ttd=""
  Direktori=namafile
  NamaFile=namafile
  Status="Aktif"
  KetSKKA=skka
  Retensi=retensi
  TglEntri= datetime.date.today()  
  Lokasi=file_link

  insert_arsip(Pengguna_id, UnitKerja_id, KA_id, Berkas_id, NomorSurat, TglSurat, Kode, Jenis, Hal, IsiSurat, Lampiran, Ttd, Direktori, NamaFile, Status, KetSKKA, Retensi, TglEntri, Lokasi, Pengguna, tahun)

  return

def check_if_file_is_open(file_path):
  try:
    with open(file_path, 'r+'):
      cek = True #jika file sdg tidak terbuka
  except IOError:
    cek = False #jika file sedang terbuka

  return cek

def is_pdf_scanned(pdf_path):    
    doc = fitz.open(stream=pdf_path.read(), filetype="pdf")
    is_scanned = True

    # Cek setiap halaman apakah hanya berisi gambar atau ada teks
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)  # Load halaman
        text = page.get_text("text")    # Ekstrak teks dari halaman
        
        if text.strip():  # Jika ada teks pada halaman
            is_scanned = False
            break

    return is_scanned

def merge_lines(text):
  lines = text.splitlines()  # Memisahkan teks menjadi baris-baris
  merged_text = ""
  for i, line in enumerate(lines):
    line = line.strip()  # Menghapus spasi atau karakter kosong di awal dan akhir baris

    # Jika baris tidak kosong dan tidak diakhiri dengan tanda akhir kalimat
    if line and not line.endswith((".", "!", "?", ":", ";")):
      merged_text += line + " "
    else:
      merged_text += line + "\n"

  return merged_text.strip()  # Menghapus spasi kosong di awal dan akhir teks hasil akhir

# Fungsi untuk mengekstrak atribut dari teks menggunakan NER (Nomor, Tanggal, Kode Surat, Hal)
def extract_attributes(text): 
    #doc = nlp(text)    
    tahun = None
    nomor_surat = "xxxxxxxxxxxxxxxxx"
    hal_surat = None
    tanggal_surat = None
    kode_surat = None
    retensi = None
    skka = None
    jumlah = 0
    baca = True

    # Ekstrak nomor surat dan hal surat dari pola teks
    nomor_surat_match = re.search(r'(?i)\b(?:No.|Nomor)(?: surat)?\s*:\s*([\w/.-]+)', text, re.IGNORECASE)
    if nomor_surat_match:
      nomor_surat = nomor_surat_match.group(0)
      nomor_surat = nomor_surat.replace(" ","")
      nomor_surat = re.sub(r'Nomor\s*:', '', nomor_surat)
      nomor_surat = re.sub(r'No\s*:', '', nomor_surat)
      nomor_surat = re.sub(r'No.\s*:', '', nomor_surat)

    tahun = nomor_surat[-4:] # mendapatkan tahun
    skka = nomor_surat[-6:] #mendapatkan skka
    skka = skka[0] #mendapatkan kode skka

    retensi = nomor_surat[-8:] 
    retensi = retensi[0] #mendapatkan kode retensi

    #mendapatkan kode surat
    kode_surat = nomor_surat[-17:]
    kode_surat = kode_surat[:8]
    posisi = kode_surat.find('/')
    if posisi == 2:
      kode_surat = kode_surat[-5:]
    elif posisi == 5:
      kode_surat = kode_surat[-2:]
    else:
      kode_surat = kode_surat

    #mendapatkan hal surat
    hal_surat_match = re.search(r'\b(?:Hal|Perihal)(?: surat)?\s*:\s*(.*)', text, re.IGNORECASE)
    if hal_surat_match:
        hal_surat = hal_surat_match.group(1)
    else:
      surtug = re.search(r'\bSurat Tugas(?: surat)?\s*\s*(.*)', text, re.IGNORECASE)
      if surtug:
        hal_surat = "Surat Tugas " 
      
      surket = re.search(r'\bSurat Keterangan(?: surat)?\s*\s*(.*)', text, re.IGNORECASE)
      if surket:
        hal_surat = "Surat Keterangan " 

      surkuasa = re.search(r'\bSurat Kuasa(?: surat)?\s*\s*(.*)', text, re.IGNORECASE)
      if surkuasa:
        hal_surat = "Surat Kuasa " 

      surpernyataan = re.search(r'\bSurat Pernyataan(?: surat)?\s*\s*(.*)', text, re.IGNORECASE)
      if surket:
        hal_surat = "Surat Pernyataan " 

      sk = re.search(r'KEPUTUSAN', text, re.IGNORECASE)
      if sk:
        hal_surat = "Surat Keputusan "
        tentang = re.search(r'Tentang :\s*(.*?)(?:\n|$)', text, re.DOTALL) 
        if tentang:
          hal_surat = hal_surat + tentang.group(0)
      #else:
      #  hal_surat = "None / tidak terbaca"

    # Menggunakan NER untuk menemukan tanggal
    # Pola untuk tanggal dengan nama bulan (baik singkat atau penuh)
    pola = r'\b\d{1,2} (Januari|Jan|Februari|Feb|Maret|Mar|April|Apr|Mei|Juni|Jun|Juli|Jul|Agustus|Agust|September|Sept|Oktober|Okt|November|Nov|Desember|Des) \d{4}?\b'

    # Mencari tanggal dalam teks
    tanggal_surat = re.search(pola, text)
    if tanggal_surat:
      tanggal_surat = tanggal_surat.group()
    else:
      tanggal_surat = datetime.date.today() 
    
    return{
      'tahun': tahun,
      'tanggal_surat': tanggal_surat,
      'nomor_surat': nomor_surat,
      'hal_surat': hal_surat,
      'kode_surat': kode_surat,
      'retensi': retensi,
      'skka': skka,
      }

# Fungsi untuk mengambil data dari database
def fetch_data_from_db():
    conn = sqlite3.connect('simarsip.db')  # Ganti dengan nama database Anda
    query = "SELECT NomorSurat, TglSurat, Kode, Hal, Lokasi, Retensi, KetSKKA, Lampiran, TglEntri  FROM ArsipIn Order by id desc"  # Ganti dengan query SQL yang sesuai
    df = pd.read_sql(query, conn)  # Membaca data dari database ke dalam DataFrame
    conn.close()  # Menutup koneksi database
    return df

def grid_data(df, rows_per_page):  
  gb = GridOptionsBuilder.from_dataframe(df)
  gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=rows_per_page)  # Atur jumlah baris per halaman
  gb.configure_side_bar()  # Mengaktifkan sidebar filter di grid
  grid_options = gb.build()
    
  AgGrid(df, gridOptions=grid_options, enable_enterprise_modules=True)

  return

def utama_view():
  rows_per_page = st.selectbox("Rows per page:", [5, 10, 15, 20], index=1)
  df = fetch_data_from_db()

  if not df.empty:
    grid_data(df, rows_per_page)  # Tampilkan data dari database
  else:
    st.error("Tidak ada data yang ditemukan di database.")

  return

def utama_download():
  if os.path.exists("DaftarArsipMasuk.xlsx"):
    os.startfile("DaftarArsipMasuk.xlsx")
  else:
    st.error("Proses download gagal / data tidak ada....")

def entri_manual():
  col1, col2 = st.columns(2)
  file_upload = col1.file_uploader("Pilih file PDF", type=["pdf"])

  #file_upload = st.file_uploader("Upload File Pdf", type=["pdf"])
  if file_upload is not None:
    st.write(f"File {file_upload.name} telah diunggah.")
    upload_folder = "uploaded_files/"
    namafile = file_upload.name
    if not os.path.exists(upload_folder):
      os.makedirs(upload_folder)
    save_path = os.path.join(upload_folder, file_upload.name)
            
    # Menyimpan file di server lokal
    with open(save_path, "wb") as f:
      f.write(file_upload.getbuffer())
    st.success("file PDF terunggah.")
  else:
    st.error("Harap unggah file PDF sebelum menyimpan data.")

  # Mengambil data untuk Klasifikasi Primer
  KAprimer = get_Primer()
  daf_primer = [f"{unit[1]} - {unit[2]}" for unit in KAprimer]  # Gabungan kode dan nama primer
  selected_primer = st.selectbox("Pilih Klasifikasi Primer", daf_primer)

  # Mengambil primer_id yang dipilih
  primer_id = [unit[1] for unit in KAprimer if f"{unit[1]} - {unit[2]}" == selected_primer][0]
  kode_surat = primer_id
  KAsekunder = get_Sekunder(primer_id)
  if KAsekunder:
    daf_sekunder = [f"{unit[1]} - {unit[2]}" for unit in KAsekunder]  # Gabungan kode dan nama sekunder
    selected_sekunder = st.selectbox("Pilih Klasifikasi Sekunder", daf_sekunder)
    sekunder_id = [unit[1] for unit in KAsekunder if f"{unit[1]} - {unit[2]}" == selected_sekunder][0]
    ketSKKA = [unit[4] for unit in KAsekunder if f"{unit[1]} - {unit[2]}" == selected_sekunder][0]
    StatusRetensi = [unit[5] for unit in KAsekunder if f"{unit[1]} - {unit[2]}" == selected_sekunder][0]
    Deskripsi = [unit[6] for unit in KAsekunder if f"{unit[1]} - {unit[2]}" == selected_sekunder][0]
    kode_surat = sekunder_id
    if Deskripsi is not None:
      st.write("Deskripsi klasifikasi arsip : ", Deskripsi)

    KATersier = get_Tersier(sekunder_id)
    if KATersier:
      daf_tersier = [f"{unit[1]} - {unit[2]}" for unit in KATersier]  # Gabungan kode dan nama sekunder
      selected_tersier = st.selectbox("Pilih Klasifikasi Tersier", daf_tersier)
      tersier_id = [unit[1] for unit in KATersier if f"{unit[1]} - {unit[2]}" == selected_tersier][0]
      ketSKKA = [unit[4] for unit in KAsekunder if f"{unit[1]} - {unit[2]}" == selected_sekunder][0]
      StatusRetensi = [unit[5] for unit in KAsekunder if f"{unit[1]} - {unit[2]}" == selected_sekunder][0]
      kode_surat = tersier_id
      Deskripsi = [unit[6] for unit in KATersier if f"{unit[1]} - {unit[2]}" == selected_tersier][0]
      if Deskripsi is not None:
        st.write("Deskripsi klasifikasi arsip : ", Deskripsi)
  else:
    st.write("Tidak ada data Klasifikasi Sekunder untuk pilihan Primer ini.")
   
  options_retensi = ["M", "P", "D"]
  if StatusRetensi in options_retensi:
    StatusRetensi = options_retensi.index(StatusRetensi)
  retensi = st.selectbox("Jenis Retensi", options_retensi, index=StatusRetensi)

  options_skka = ["B", "T"]
  if ketSKKA in options_skka:
    ketSKKA = options_skka.index(ketSKKA)
  skka = st.selectbox("Kode SKKA", options_skka, index=ketSKKA)

  no = st.number_input("Nomor Surat", min_value=1, step=1)
  kodeunit = st.text_input("Kode Unit Kerja Pengirim")
  strtanggal_surat = st.date_input("Tanggal Surat", value=datetime.date.today())
  tahun = strtanggal_surat.year
  hal_surat = st.text_input("Perihal Surat")
  jum_hal = st.number_input("Jumlah Lampiran", min_value=0, step=1)
  tanggal_surat = strtanggal_surat.strftime("%d/%b/%Y")
  nomor_surat = f"{no}/{kodeunit}/{kode_surat}/{retensi}/{skka}/{tahun}"

  if st.button("Simpan Data"):    
    berkas(kode_surat, save_path, namafile, tahun, nomor_surat, tanggal_surat,  hal_surat,  retensi, skka, jum_hal)
  return

def utama_uploadIn():
  #st.subheader("Upload Arsip Masuk")
  pdf_document = ""
  with st.container(): 
    col1, col2, col3= st.columns([5,2,2]) #[8,3,2]

    Upload_button = col1.file_uploader("Entri Arsip Masuk dengan Upload File Pdf", type=["pdf"])
    #download_button = col3.button("Download")
    if Upload_button:
      pdf_document = Upload_button
      if is_pdf_scanned(pdf_document) == False:
        pdf_document.seek(0)  # Reset pointer ke awal
        convert_pdfNonscan_text(pdf_document, pdf_document.name)
      else:
        try:
          pdf_document.seek(0)  # Reset pointer ke awal
          convert_pdfscan_text1(pdf_document, pdf_document.name)
        except:
          pdf_document.seek(0)  # Reset pointer ke awal
          convert_pdfscan_text2(pdf_document, pdf_document.name)
      
    utama_view()
    explor.explor_arsip()
    surat_masuk = get_aktif()
    if surat_masuk:
      edit_retensi()
      df = pd.DataFrame(surat_masuk, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInaktif", "ThnMusnah_Serah", "Lokasi"])
      excel_data = convert_df_to_excel(df)
      st.download_button(
        label="Download file Excel",
        data=excel_data,
        file_name="arsip_masuk.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
      )

    #if st.button("Download"):
    #  utama_download()

  return

