
import os
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import streamlit as st
import sqlite3
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime
import io

def grid_aktif(df, rows_per_page):  
  df = df[['Tahun', 'NomorSurat', 'TglSurat', 'Kode', 'Hal', 'Lampiran', 'KetSKKA', 'Retensi', 'RetAktif', 'RetInaktif', 'ThnInaktif', 'ThnMusnah_Serah', 'Lokasi', 'Status']]   
  gb = GridOptionsBuilder.from_dataframe(df)
  gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=rows_per_page)  # Atur jumlah baris per halaman
  gb.configure_side_bar()  # Mengaktifkan sidebar filter di grid
  grid_options = gb.build()
    
  AgGrid(df, gridOptions=grid_options, enable_enterprise_modules=True)

  return

def grid_inaktif(df, rows_per_page):  
  df = df[['Tahun', 'NomorSurat', 'TglSurat', 'Kode', 'Hal', 'Lampiran', 'KetSKKA', 'Retensi', 'RetAktif', 'RetInaktif', 'ThnInAktif', 'ThnMusnah_Serah', 'Lokasi', 'Status']]   
  gb = GridOptionsBuilder.from_dataframe(df)
  gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=rows_per_page)  # Atur jumlah baris per halaman  ThnInAktif, ThnMusnah_Serah,
  gb.configure_side_bar()  # Mengaktifkan sidebar filter di grid
  grid_options = gb.build()
    
  AgGrid(df, gridOptions=grid_options, enable_enterprise_modules=True)

  return

def edit_retensi():
    conn = sqlite3.connect('simarsip.db')
    cursor = conn.cursor()

    query = """
    UPDATE ArsipOut
    SET Retensi = (SELECT KA.StatusRetensi FROM KA WHERE KA.Kode = ArsipOut.Kode),
        KetSKKA = (SELECT KA.KetSKKA FROM KA WHERE KA.Kode = ArsipOut.Kode),
        RetAktif = (SELECT KA.RetAktif FROM KA WHERE KA.Kode = ArsipOut.Kode),
        Retinaktif = (SELECT KA.RetInaktif FROM KA WHERE KA.Kode = ArsipOut.Kode)
    WHERE EXISTS (SELECT 1 FROM KA WHERE KA.Kode = ArsipOut.Kode);
    """
    cursor.execute(query)
    conn.commit()

    query = """
    UPDATE ArsipIn
    SET Retensi = (SELECT KA.StatusRetensi FROM KA WHERE KA.Kode = ArsipIn.Kode),
        KetSKKA = (SELECT KA.KetSKKA FROM KA WHERE KA.Kode = ArsipIn.Kode),
        RetAktif = (SELECT KA.RetAktif FROM KA WHERE KA.Kode = ArsipIn.Kode),
        Retinaktif = (SELECT KA.RetInaktif FROM KA WHERE KA.Kode = ArsipIn.Kode)
    WHERE EXISTS (SELECT 1 FROM KA WHERE KA.Kode = ArsipIn.Kode);
    """
    cursor.execute(query)
    conn.commit()

    conn.close()

    return 

def get_aktif():
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  query="""
  SELECT  Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInaktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipout WHERE Status='Aktif' 
  UNION 
  SELECT  Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInaktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipin WHERE Status='Aktif' 
  order by Kode, TglSurat, NomorSurat
  """
  cursor.execute(query)  
  aktif = cursor.fetchall()
  conn.close()
  return aktif

def get_inaktif(thn):
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  query="""
  SELECT Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInAktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipout WHERE RetAktif+1+Tahun=? 
  UNION 
  SELECT Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInAktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipin WHERE RetAktif+1+Tahun=?
  order by Kode, TglSurat, NomorSurat
  """
  cursor.execute(query, (thn, thn))  
  inaktif = cursor.fetchall()
  conn.close()
  return inaktif

def get_musnah(thn):
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  query="""
  SELECT Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInAktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipout WHERE Retensi='M' AND RetAktif+RetInaktif+1+Tahun=? 
  UNION 
  SELECT Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInAktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipin WHERE Retensi='M' AND RetAktif+RetInaktif+1+Tahun=?
  order by Kode, TglSurat, NomorSurat
  """
  cursor.execute(query, (thn, thn))  
  musnah = cursor.fetchall()
  conn.close()
  return musnah

def get_permanen(thn):
  conn = sqlite3.connect('simarsip.db')
  cursor = conn.cursor()
  query="""
  SELECT Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInAktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipout WHERE Retensi='P' AND RetAktif+RetInaktif+1+Tahun=? 
  UNION 
  SELECT Tahun, NomorSurat, TglSurat, Kode, Hal, Lampiran, KetSKKA, Retensi, RetAktif, RetInaktif, RetAktif+1+Tahun AS ThnInAktif, RetAktif+RetInaktif+1+Tahun AS ThnMusnah_Serah, Lokasi, Status FROM Arsipin WHERE Retensi='P' AND RetAktif+RetInaktif+1+Tahun=?
  order by Kode, TglSurat, NomorSurat
  """
  cursor.execute(query, (thn, thn))  
  musnah = cursor.fetchall()
  conn.close()
  return musnah

bulan_mapping = {
    "Januari": "01", "Februari": "02", "Maret": "03", "April": "04", "Mei": "05", "Juni": "06",
    "Juli": "07", "Agustus": "08", "September": "09", "Oktober": "10", "November": "11", "Desember": "12"
}

def convert_to_date(date_str):
    try:
        # Pisahkan tanggal menjadi bagian-bagian (tanggal, bulan, tahun)
        parts = date_str.split()
        if len(parts) == 3:
            day = parts[0]
            month = bulan_mapping.get(parts[1], "01")  # Ambil bulan dari kamus
            year = parts[2]
            # Gabungkan kembali dalam format yang bisa di-parse oleh datetime
            date_obj = datetime.strptime(f"{day}/{month}/{year}", "%d/%m/%Y")
            return date_obj.year  # Kembalikan tahun
        else:
            return None  # Jika format tidak sesuai
    except Exception as e:
        return None  # Jika ada kesalahan parsing

def convert_df_to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        # Mengakses workbook dan worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Membuat hyperlink pada kolom "Lokasi" (misalnya kolom G)
        for idx, link in enumerate(df['Lokasi'], start=1):
            worksheet.write_url(f'M{idx + 1}', link, string=link)

    processed_data = output.getvalue()
    return processed_data

def lap_aktif():
    edit_retensi()
    data_aktif = get_aktif()
    rows_per_page = st.selectbox("Rows per page:", [5, 10, 15, 20], index=1)
    df = pd.DataFrame(data_aktif, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", 'RetInaktif', 'ThnInaktif', 'ThnMusnah_Serah', "Lokasi", "Status"])
    #df['Tahun'] = df['TglSurat'].apply(convert_to_date)
    #df = fetch_data_from_db()

    if not df.empty:
      grid_aktif(df, rows_per_page)  # Tampilkan data dari database
    else:
      st.error("Tidak ada data yang ditemukan di database.")

    file_name = "laporan_arsip_aktif.xlsx"
    excel_data = convert_df_to_excel(df)
    st.download_button(
      label="Unduh Daftar Arsip Aktif",
      data=excel_data,
      file_name="laporan_arsip_aktif.xlsx",
      mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

        #df.to_excel(file_name, index=False, engine='openpyxl')

        # Memuat workbook Excel yang baru dibuat
        #wb = load_workbook(file_name)
        #ws = wb.active            
        # Menambahkan hyperlink ke kolom Lokasi
        #for row in range(2, ws.max_row + 1):  # Mulai dari baris 2 (karena baris 1 adalah header)
        #    cell = ws[f'M{row}']  # Kolom H adalah kolom Lokasi
        #    location_url = cell.value
        #    if location_url:
        #        cell.hyperlink = location_url  # Set hyperlink
        #        cell.font = Font(color="0000FF", underline="single")  # Format sebagai link (biru dan underline)

        # Simpan ulang workbook
        #wb.save(file_name)
        
        #print(f"Laporan arsip aktif telah disimpan di {file_name}")
        #if os.path.exists("laporan_arsip_aktif.xlsx"):
        #  os.startfile("laporan_arsip_aktif.xlsx")
        #else:
        #  st.error("Proses download gagal / data tidak ada....")
  
    return

def lap_inaktif():
    edit_retensi()
    thn_skrg = datetime.now().year
    list_thn = list(range(thn_skrg, thn_skrg+5))
    thn = st.selectbox("Pilih Daftar Arsip Usul Pindah pada Tahun ", list_thn)

    data_inaktif = get_inaktif(thn)     
    rows_per_page = st.selectbox("Rows per page:", [5, 10, 15, 20], index=1)
    df = pd.DataFrame(data_inaktif, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInAktif", "ThnMusnah_Serah", "Lokasi", "Status"]) 

    if not df.empty:
      grid_inaktif(df, rows_per_page)  # Tampilkan data dari database
    else:
      st.error("Tidak ada data yang ditemukan di database.")

    if st.button("Unduh Daftar Arsip InAktif"):  
      if data_inaktif:
        df = pd.DataFrame(data_inaktif, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInaktif", "ThnMusnah_Serah", "Lokasi", "Status"])
            
        file_name = "laporan_arsip_inaktif.xlsx"
        df.to_excel(file_name, index=False, engine='openpyxl')

        # Memuat workbook Excel yang baru dibuat
        wb = load_workbook(file_name)
        ws = wb.active            
        # Menambahkan hyperlink ke kolom Lokasi
        for row in range(2, ws.max_row + 1):  # Mulai dari baris 2 (karena baris 1 adalah header)
            cell = ws[f'M{row}']  # Kolom H adalah kolom Lokasi
            location_url = cell.value
            if location_url:
                cell.hyperlink = location_url  # Set hyperlink
                cell.font = Font(color="0000FF", underline="single")  # Format sebagai link (biru dan underline)

        # Simpan ulang workbook
        wb.save(file_name)
        
        #print(f"Laporan arsip aktif telah disimpan di {file_name}")
        if os.path.exists("laporan_arsip_inaktif.xlsx"):
          os.startfile("laporan_arsip_inaktif.xlsx")
        else:
          st.error("Proses download gagal / data tidak ada....")
      else:
        print("Tidak ada arsip aktif yang ditemukan.")
  
    return

def lap_musnah():
    edit_retensi()
    thn_skrg = datetime.now().year
    list_thn = list(range(thn_skrg, thn_skrg+5))
    thn = st.selectbox("Pilih Daftar Arsip Usul Musnah pada Tahun ", list_thn)

    data_musnah = get_musnah(thn)     
    rows_per_page = st.selectbox("Rows per page:", [5, 10, 15, 20], index=1)
    df = pd.DataFrame(data_musnah, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInAktif", "ThnMusnah_Serah", "Lokasi", "Status"]) 

    if not df.empty:
      grid_inaktif(df, rows_per_page)  # Tampilkan data dari database
    else:
      st.error("Tidak ada data yang ditemukan di database.")

    if st.button("Unduh Daftar Arsip Musnah"):  
      if data_musnah:
        df = pd.DataFrame(data_musnah, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInaktif", "ThnMusnah_Serah", "Lokasi", "Status"])
            
        file_name = "laporan_arsip_musnah.xlsx"
        df.to_excel(file_name, index=False, engine='openpyxl')

        # Memuat workbook Excel yang baru dibuat
        wb = load_workbook(file_name)
        ws = wb.active            
        # Menambahkan hyperlink ke kolom Lokasi
        for row in range(2, ws.max_row + 1):  # Mulai dari baris 2 (karena baris 1 adalah header)
            cell = ws[f'M{row}']  # Kolom H adalah kolom Lokasi
            location_url = cell.value
            if location_url:
                cell.hyperlink = location_url  # Set hyperlink
                cell.font = Font(color="0000FF", underline="single")  # Format sebagai link (biru dan underline)

        # Simpan ulang workbook
        wb.save(file_name)
        
        #print(f"Laporan arsip aktif telah disimpan di {file_name}")
        if os.path.exists("laporan_arsip_musnah.xlsx"):
          os.startfile("laporan_arsip_musnah.xlsx")
        else:
          st.error("Proses download gagal / data tidak ada....")
      else:
        print("Tidak ada arsip aktif yang ditemukan.")
  
    return

def lap_statis():
    edit_retensi()
    thn_skrg = datetime.now().year
    list_thn = list(range(thn_skrg, thn_skrg+5))
    thn = st.selectbox("Pilih Daftar Arsip Usul Serah pada Tahun ", list_thn)

    data_permanen = get_permanen(thn)     
    rows_per_page = st.selectbox("Rows per page:", [5, 10, 15, 20], index=1)
    df = pd.DataFrame(data_permanen, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInAktif", "ThnMusnah_Serah", "Lokasi", "Status"]) 

    if not df.empty:
      grid_inaktif(df, rows_per_page)  # Tampilkan data dari database
    else:
      st.error("Tidak ada data yang ditemukan di database.")

    if st.button("Unduh Daftar Arsip Statis"):  
      if data_permanen:
        df = pd.DataFrame(data_permanen, columns=["Tahun", "NomorSurat", "TglSurat", "Kode", "Hal", "Lampiran", "KetSKKA", "Retensi", "RetAktif", "RetInaktif", "ThnInaktif", "ThnMusnah_Serah", "Lokasi", "Status"])
            
        file_name = "laporan_arsip_statis.xlsx"
        df.to_excel(file_name, index=False, engine='openpyxl')

        # Memuat workbook Excel yang baru dibuat
        wb = load_workbook(file_name)
        ws = wb.active            
        # Menambahkan hyperlink ke kolom Lokasi
        for row in range(2, ws.max_row + 1):  # Mulai dari baris 2 (karena baris 1 adalah header)
            cell = ws[f'M{row}']  # Kolom H adalah kolom Lokasi
            location_url = cell.value
            if location_url:
                cell.hyperlink = location_url  # Set hyperlink
                cell.font = Font(color="0000FF", underline="single")  # Format sebagai link (biru dan underline)

        # Simpan ulang workbook
        wb.save(file_name)
        
        #print(f"Laporan arsip aktif telah disimpan di {file_name}")
        if os.path.exists("laporan_arsip_statis.xlsx"):
          os.startfile("laporan_arsip_statis.xlsx")
        else:
          st.error("Proses download gagal / data tidak ada....")
      else:
        print("Tidak ada arsip aktif yang ditemukan.")
    return

